from rest_framework import permissions, serializers, status, viewsets
from rest_framework.decorators import action
from rest_framework.response import Response

from users.views import TenantQuerySetMixin
from users.permissions import ActionBasedPermission

from .models import Customer, CustomerContact, CustomerInteraction, CustomerTag
from .serializers import (
    CustomerContactSerializer,
    CustomerInteractionSerializer,
    CustomerSerializer,
    CustomerTagSerializer,
)


class CustomerTagViewSet(TenantQuerySetMixin, viewsets.ModelViewSet):
    """CRUD Tags khách hàng — cô lập theo company."""
    module_code = "crm"

    queryset = CustomerTag.objects.select_related("company").order_by("name")
    serializer_class = CustomerTagSerializer
    permission_classes = [permissions.IsAuthenticated, ActionBasedPermission]
    
    action_permissions = {
        "list": "crm.manage_tags",
        "retrieve": "crm.manage_tags",
        "create": "crm.manage_tags",
        "update": "crm.manage_tags",
        "partial_update": "crm.manage_tags",
        "destroy": "crm.manage_tags",
    }


class CustomerViewSet(TenantQuerySetMixin, viewsets.ModelViewSet):
    """CRUD khách hàng — luôn filter theo company của user đang đăng nhập."""
    module_code = "crm"

    queryset = Customer.objects.select_related(
        "company", "assigned_to", "created_by"
    ).prefetch_related("contacts", "interactions", "tags").order_by("-created_at")
    serializer_class = CustomerSerializer
    permission_classes = [permissions.IsAuthenticated, ActionBasedPermission]
    
    action_permissions = {
        "list": "crm.view",
        "retrieve": "crm.view",
        "create": "crm.create",
        "update": "crm.edit",
        "partial_update": "crm.edit",
        "destroy": "crm.delete",
        "assign": "crm.assign",
        "round_robin_assign": "crm.auto_assign",
        "import_excel": "crm.import",
    }

    def get_queryset(self):
        qs = super().get_queryset()
        user = self.request.user
        # Phân quyền xem dữ liệu
        if not user.is_company_admin and not user.is_superuser and not user.has_perm_code("crm.view_all"):
            managed_deps = user.managed_departments.all()
            if managed_deps.exists():
                from django.db.models import Q, Count
                qs = qs.filter(
                    Q(assigned_to=user) | 
                    Q(assigned_to__department__in=managed_deps)
                )
            else:
                qs = qs.filter(assigned_to=user)
        # Filter theo trạng thái nếu có query param
        customer_status = self.request.query_params.get("status")
        if customer_status:
            qs = qs.filter(status=customer_status)
        # Filter theo assigned_to nếu có query param (dành cho manager)
        assigned_to = self.request.query_params.get("assigned_to")
        if assigned_to:
            qs = qs.filter(assigned_to_id=assigned_to)
        # Tìm kiếm theo tên hoặc SĐT
        search = self.request.query_params.get("search")
        if search:
            from django.db.models import Q
            qs = qs.filter(Q(name__icontains=search) | Q(phone__icontains=search))
            
        from django.db.models import Count
        qs = qs.annotate(
            quotation_count=Count('quotations', distinct=True),
            order_count=Count('orders', distinct=True)
        )
        return qs

    def perform_create(self, serializer):
        company = self.request.user.company
        user = self.request.user
        default_assignee = user if user.has_perm_code("crm.auto_assign_self") else None
        assigned_to = serializer.validated_data.get('assigned_to', default_assignee)
        serializer.save(company=company, created_by=user, assigned_to=assigned_to)

    def destroy(self, request, *args, **kwargs):
        from django.db.models.deletion import ProtectedError
        try:
            return super().destroy(request, *args, **kwargs)
        except ProtectedError:
            return Response(
                {"detail": "Không thể xóa khách hàng này vì đã có dữ liệu liên quan (Báo giá, Đơn hàng...)."},
                status=status.HTTP_400_BAD_REQUEST
            )

    @action(detail=True, methods=["post"], url_path="assign")
    def assign(self, request, pk=None):
        """
        POST /api/crm/customers/{id}/assign/
        Gán thủ công một Sale cụ thể cho khách hàng.
        Yêu cầu quyền: crm.assign hoặc company admin.
        Body: { "assigned_to": <user_id> }
        """
        customer = self.get_object()

        if not request.user.is_company_admin and not request.user.has_perm_code("crm.assign"):
            return Response(
                {"detail": "Bạn không có quyền phân công khách hàng."},
                status=status.HTTP_403_FORBIDDEN,
            )

        assigned_to_id = request.data.get("assigned_to")
        if not assigned_to_id:
            return Response(
                {"detail": "Vui lòng cung cấp 'assigned_to' (ID nhân viên)."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        from users.models import User
        try:
            sale_user = User.objects.get(
                id=assigned_to_id,
                company=request.user.company,
                is_active=True,
            )
        except User.DoesNotExist:
            return Response(
                {"detail": "Nhân viên không tồn tại hoặc không thuộc công ty này."},
                status=status.HTTP_404_NOT_FOUND,
            )

        customer._assigned_by = request.user  # Signal sẽ đọc để biết ai gán
        customer.assigned_to = sale_user
        customer.save(update_fields=["assigned_to", "updated_at"])

        # Reload từ DB để có đủ select_related
        customer = Customer.objects.select_related("assigned_to", "company", "created_by").get(pk=customer.pk)
        return Response(CustomerSerializer(customer, context={"request": request}).data)

    @action(detail=False, methods=["post"], url_path="round-robin-assign")
    def round_robin_assign(self, request):
        """
        POST /api/crm/customers/round-robin-assign/
        Phân bổ tự động khách hàng chưa có nhân viên theo Round-robin.
        Chỉ Company Admin mới gọi được.
        """
        if not request.user.is_company_admin and not request.user.has_perm_code("crm.auto_assign"):
            return Response(
                {"detail": "Bạn không có quyền thực hiện Phân bổ khách hàng tự động."},
                status=status.HTTP_403_FORBIDDEN,
            )

        from users.models import User
        from django.db.models import Max, F

        company = request.user.company

        # Lấy danh sách Sale đang hoạt động, sắp xếp theo thời gian nhận Lead gần nhất (ai chờ lâu nhất / mới vào thì ưu tiên trước)
        sale_users = list(
            User.objects.filter(
                company=company,
                is_active=True,
                is_company_admin=False,
                is_superuser=False,
                role__is_auto_assign_target=True,
            ).annotate(
                last_lead_time=Max("assigned_customers__created_at")
            ).order_by(F("last_lead_time").asc(nulls_first=True))
        )

        if not sale_users:
            return Response(
                {"detail": "Không có nhân viên nào thuộc nhóm được chia khách tự động."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        unassigned = Customer.objects.filter(
            company=company,
            assigned_to__isnull=True,
        )

        assigned_count = 0
        for i, customer in enumerate(unassigned):
            sale = sale_users[i % len(sale_users)]
            customer._assigned_by = request.user
            customer.assigned_to = sale
            customer.save(update_fields=["assigned_to", "updated_at"])
            assigned_count += 1

        return Response(
            {"detail": f"Đã phân bổ tự động thành công {assigned_count} khách hàng."},
            status=status.HTTP_200_OK,
        )

    @action(detail=False, methods=["get"], url_path="export-csv")
    def export_csv(self, request):
        """
        GET /api/crm/customers/export-csv/
        Xuất danh sách khách hàng ra file Excel (.xlsx).
        """
        import openpyxl
        from django.http import HttpResponse

        qs = self.get_queryset()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "KhachHang"
        ws.append([
            'Họ và tên', 'Số điện thoại', 'Email', 'Địa chỉ', 'Tỉnh/Thành phố', 'Tags',
            'Nguồn khách', 'Trạng thái', 'Nhân viên phụ trách', 'Ngày tạo'
        ])

        for customer in qs:
            assigned = customer.assigned_to.full_name if customer.assigned_to else ''
            tags_str = ", ".join(t.name for t in customer.tags.all())
            ws.append([
                customer.name,
                customer.phone,
                customer.email,
                customer.address,
                customer.city,
                tags_str,
                customer.get_source_display(),
                customer.get_status_display(),
                assigned,
                customer.created_at.strftime('%Y-%m-%d %H:%M') if customer.created_at else ''
            ])

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="customers.xlsx"'
        wb.save(response)
        return response

    @action(detail=False, methods=["get"], url_path="export-template")
    def export_template(self, request):
        import openpyxl
        from django.http import HttpResponse

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "KhachHang_Mau"
        ws.append([
            'Họ và tên', 'Số điện thoại', 'Email', 'Địa chỉ', 'Tỉnh/Thành phố', 'Tags'
        ])
        ws.append([
            'Nguyễn Văn A', '0901234567', 'nguyenvana@gmail.com', '123 Lê Lợi', 'TP.HCM', 'VIP, Khách sỉ'
        ])
        ws.append([
            'Công ty TNHH B', '0987654321', '', '', '', 'Khách mới'
        ])

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="mau_nhap_khach_hang.xlsx"'
        wb.save(response)
        return response

    @action(detail=False, methods=["post"], url_path="import-csv")
    def import_csv(self, request):
        """
        POST /api/crm/customers/import-csv/
        Nhập danh sách khách hàng từ file CSV.
        """
        import csv
        import io
        from rest_framework.parsers import MultiPartParser

        # Tạm thời cấu hình parser cho endpoint này
        self.parser_classes = [MultiPartParser]

        file_obj = request.FILES.get('file')
        if not file_obj:
            return Response({"detail": "Vui lòng chọn file CSV."}, status=status.HTTP_400_BAD_REQUEST)

        if not (file_obj.name.endswith('.csv') or file_obj.name.endswith('.xlsx')):
            return Response({"detail": "Chỉ hỗ trợ định dạng .csv hoặc .xlsx"}, status=status.HTTP_400_BAD_REQUEST)

        company = request.user.company
        created_count = 0
        error_count = 0
        errors = []

        try:
            if file_obj.name.endswith('.xlsx'):
                import openpyxl
                wb = openpyxl.load_workbook(file_obj, data_only=True)
                ws = wb.active
                rows = list(ws.iter_rows(values_only=True))
                if not rows or len(rows) < 2:
                    return Response({"detail": "File trống."}, status=status.HTTP_400_BAD_REQUEST)
                data_rows = rows[1:]
                
                def get_val(r, idx):
                    if idx < len(r) and r[idx] is not None:
                        return str(r[idx]).strip()
                    return ""
            else:
                decoded_file = file_obj.read().decode('utf-8-sig')
                io_string = io.StringIO(decoded_file)
                reader = csv.reader(io_string)
                headers = next(reader, None)
                if headers and headers[0].startswith('sep='):
                    headers = next(reader, None)
                if not headers:
                    return Response({"detail": "File trống."}, status=status.HTTP_400_BAD_REQUEST)
                data_rows = list(reader)
                
                def get_val(r, idx):
                    if idx < len(r):
                        return str(r[idx]).strip()
                    return ""

            # Dự kiến định dạng cột tối thiểu: Tên, SĐT. Tùy chọn: Email, Địa chỉ, Thành phố, Tags
            for i, row in enumerate(data_rows):
                if not row or not any(row): continue
                try:
                    name = get_val(row, 0)
                    phone = get_val(row, 1)
                    email = get_val(row, 2)
                    address = get_val(row, 3)
                    city = get_val(row, 4)
                    tags_str = get_val(row, 5)

                    if not name or not phone:
                        error_count += 1
                        errors.append(f"Dòng {i+2}: Thiếu Tên hoặc Số điện thoại.")
                        continue
                    
                    # Fix lỗi Excel làm mất số 0 ở đầu (hoặc số bắt đầu bằng 84)
                    phone = ''.join(filter(str.isdigit, phone))
                    if phone.startswith('84'):
                        phone = '0' + phone[2:]
                    elif len(phone) == 9 and not phone.startswith('0'):
                        phone = '0' + phone
                    
                    user = request.user
                    default_assignee = user if user.has_perm_code("crm.auto_assign_self") else None
                    
                    customer = Customer.objects.create(
                        company=company,
                        name=name,
                        phone=phone,
                        email=email,
                        address=address,
                        city=city,
                        created_by=user,
                        assigned_to=default_assignee,
                        source='other',
                        status='new'
                    )
                    
                    if tags_str:
                        tag_names = [t.strip() for t in tags_str.split(',') if t.strip()]
                        for t_name in tag_names:
                            tag, _ = CustomerTag.objects.get_or_create(company=company, name=t_name)
                            customer.tags.add(tag)

                    created_count += 1
                except Exception as e:
                    error_count += 1
                    errors.append(f"Dòng {i+2}: Lỗi xử lý ({str(e)})")

        except Exception as e:
            return Response({"detail": f"Lỗi đọc file: {str(e)}"}, status=status.HTTP_400_BAD_REQUEST)

        return Response({
            "detail": f"Đã nhập thành công {created_count} khách hàng. Lỗi: {error_count}",
            "errors": errors
        }, status=status.HTTP_200_OK)


class CustomerContactViewSet(viewsets.ModelViewSet):
    """CRUD đầu mối liên hệ — filter qua customer.company."""
    module_code = "crm"

    queryset = CustomerContact.objects.select_related("customer__company").order_by("name")
    serializer_class = CustomerContactSerializer
    permission_classes = [permissions.IsAuthenticated, ActionBasedPermission]
    
    action_permissions = {
        "list": "crm.view",
        "retrieve": "crm.view",
        "create": "crm.edit",  # Creating contact requires edit customer permission
        "update": "crm.edit",
        "partial_update": "crm.edit",
        "destroy": "crm.edit",
    }

    def get_queryset(self):
        user = self.request.user
        if user.is_superuser and user.company_id is None:
            return super().get_queryset()
        qs = self.queryset.filter(customer__company=user.company)
        
        if not user.is_company_admin and not user.is_superuser and not user.has_perm_code("crm.view_all"):
            managed_deps = user.managed_departments.all()
            if managed_deps.exists():
                from django.db.models import Q
                qs = qs.filter(
                    Q(customer__assigned_to=user) | 
                    Q(customer__assigned_to__department__in=managed_deps)
                )
            else:
                qs = qs.filter(customer__assigned_to=user)

        customer_id = self.request.query_params.get("customer_id")
        if customer_id:
            qs = qs.filter(customer_id=customer_id)
        return qs


class CustomerInteractionViewSet(viewsets.ModelViewSet):
    """CRUD lịch sử chăm sóc — filter qua customer.company."""
    module_code = "crm"

    queryset = CustomerInteraction.objects.select_related(
        "customer__company", "created_by"
    ).order_by("-created_at")
    serializer_class = CustomerInteractionSerializer
    permission_classes = [permissions.IsAuthenticated, ActionBasedPermission]
    
    action_permissions = {
        "list": "crm.view",
        "retrieve": "crm.view",
        "create": "crm.edit",  # Creating interaction requires edit customer permission
        "update": "crm.edit",
        "partial_update": "crm.edit",
        "destroy": "crm.edit",
    }

    def get_queryset(self):
        user = self.request.user
        if user.is_superuser and user.company_id is None:
            return super().get_queryset()
        qs = self.queryset.filter(customer__company=user.company)

        if not user.is_company_admin and not user.is_superuser and not user.has_perm_code("crm.view_all"):
            managed_deps = user.managed_departments.all()
            if managed_deps.exists():
                from django.db.models import Q
                qs = qs.filter(
                    Q(customer__assigned_to=user) | 
                    Q(customer__assigned_to__department__in=managed_deps)
                )
            else:
                qs = qs.filter(customer__assigned_to=user)

        # Lọc theo customer nếu có query param
        customer_id = self.request.query_params.get("customer_id")
        if customer_id:
            qs = qs.filter(customer_id=customer_id)
        return qs

    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user)

    @action(detail=True, methods=["post"], url_path="upload-files")
    def upload_files(self, request, pk=None):
        """Upload nhiều file cho một tương tác (Lịch sử chăm sóc)."""
        if not request.user.has_perm_code("crm.upload_interaction_files"):
            return Response({"detail": "Bạn không có quyền tải file lên lịch sử chăm sóc."}, status=status.HTTP_403_FORBIDDEN)
            
        interaction = self.get_object()
        
        if 'files' not in request.FILES:
            return Response({"detail": "Không tìm thấy file nào để tải lên."}, status=status.HTTP_400_BAD_REQUEST)
            
        files = request.FILES.getlist('files')
        
        from .models import InteractionAttachment
        from .serializers import InteractionAttachmentSerializer
        
        attachments = []
        for file in files:
            attachment = InteractionAttachment.objects.create(
                interaction=interaction,
                file=file,
                file_name=file.name,
                file_size=file.size
            )
            attachments.append(attachment)
            
        return Response(
            InteractionAttachmentSerializer(attachments, many=True).data,
            status=status.HTTP_201_CREATED
        )
