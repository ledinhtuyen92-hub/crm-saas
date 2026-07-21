from rest_framework import permissions, viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from django.db import transaction

from users.views import TenantQuerySetMixin
from users.permissions import ActionBasedPermission

from .models import (
    InventoryTransaction, Product, ProductCategory, StockLevel, Warehouse,
    ProductTemplate, ProductAttribute, ProductAttributeValue
)
from .serializers import (
    InventoryTransactionSerializer,
    ProductCategorySerializer,
    ProductSerializer,
    StockLevelSerializer,
    WarehouseSerializer,
    ProductTemplateSerializer,
    ProductAttributeSerializer,
    ProductAttributeValueSerializer,
)


class ProductCategoryViewSet(TenantQuerySetMixin, viewsets.ModelViewSet):
    """CRUD loại sản phẩm — cô lập theo company."""
    module_code = "inventory"

    queryset = ProductCategory.objects.select_related("company").order_by("name")
    serializer_class = ProductCategorySerializer
    permission_classes = [permissions.IsAuthenticated, ActionBasedPermission]
    
    action_permissions = {
        "list": "products.manage_categories",
        "retrieve": "products.manage_categories",
        "create": "products.manage_categories",
        "update": "products.manage_categories",
        "partial_update": "products.manage_categories",
        "destroy": "products.manage_categories",
    }


class ProductTemplateViewSet(TenantQuerySetMixin, viewsets.ModelViewSet):
    """CRUD Mẫu sản phẩm."""
    module_code = "inventory"
    queryset = ProductTemplate.objects.select_related("company", "category").order_by("name")
    serializer_class = ProductTemplateSerializer
    permission_classes = [permissions.IsAuthenticated, ActionBasedPermission]
    
    action_permissions = {
        "list": ["products.view", "sales.view"],
        "retrieve": ["products.view", "sales.view"],
        "create": "products.create",
        "update": "products.edit",
        "partial_update": "products.edit",
        "destroy": "products.delete",
        "generate_variants": "products.create",
    }

    @action(detail=True, methods=['post'])
    def generate_variants(self, request, pk=None):
        template = self.get_object()
        attributes_data = request.data.get("attributes", [])
        # attributes_data format: [{"name": "Màu sắc", "values": ["Đỏ", "Xanh"]}, {"name": "Size", "values": ["M", "L"]}]
        
        if not attributes_data:
            # Generate 1 default variant
            Product.objects.get_or_create(
                company=template.company,
                template=template,
                sku=f"{template.id}-DEFAULT",
                defaults={
                    "name": template.name,
                    "category": template.category,
                    "price": 0,
                    "cost_price": 0,
                }
            )
            return Response({"detail": "Đã tạo 1 biến thể mặc định."})
        
        import itertools
        
        # Build list of value lists
        keys = []
        value_lists = []
        for attr in attributes_data:
            keys.append(attr["name"])
            value_lists.append(attr["values"])
            
        combinations = list(itertools.product(*value_lists))
        
        created_count = 0
        for combo in combinations:
            attr_dict = dict(zip(keys, combo))
            
            # sku like TEMPLATEID-DO-M
            sku_suffix = "-".join([str(v).upper() for v in combo])
            sku = f"{template.id}-{sku_suffix}"
            variant_name = f"{template.name} ({', '.join(combo)})"
            
            # Check if variant exists
            if not Product.objects.filter(company=template.company, template=template, attributes=attr_dict).exists():
                Product.objects.create(
                    company=template.company,
                    template=template,
                    sku=sku,
                    name=variant_name,
                    category=template.category,
                    attributes=attr_dict,
                    price=0,
                    cost_price=0,
                )
                created_count += 1
                
        return Response({"detail": f"Đã sinh {created_count} biến thể mới."})


class ProductAttributeViewSet(TenantQuerySetMixin, viewsets.ModelViewSet):
    """CRUD Thuộc tính sản phẩm."""
    module_code = "inventory"
    queryset = ProductAttribute.objects.select_related("company").prefetch_related("values").order_by("name")
    serializer_class = ProductAttributeSerializer
    permission_classes = [permissions.IsAuthenticated, ActionBasedPermission]
    
    action_permissions = {
        "list": ["products.view", "sales.view"],
        "retrieve": ["products.view", "sales.view"],
        "create": "products.create",
        "update": "products.edit",
        "partial_update": "products.edit",
        "destroy": "products.delete",
    }


class ProductAttributeValueViewSet(viewsets.ModelViewSet):
    """CRUD Giá trị thuộc tính sản phẩm. Không kế thừa TenantQuerySetMixin vì nó query theo attribute."""
    queryset = ProductAttributeValue.objects.select_related("attribute").order_by("value")
    serializer_class = ProductAttributeValueSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        qs = super().get_queryset()
        attribute_id = self.request.query_params.get("attribute_id")
        if attribute_id:
            qs = qs.filter(attribute_id=attribute_id)
        return qs


class ProductViewSet(TenantQuerySetMixin, viewsets.ModelViewSet):
    """CRUD sản phẩm — cô lập theo company."""
    module_code = "inventory"

    queryset = Product.objects.select_related("company", "category").order_by("name")
    serializer_class = ProductSerializer
    permission_classes = [permissions.IsAuthenticated, ActionBasedPermission]
    pagination_class = None
    
    action_permissions = {
        "list": ["products.view", "sales.view", "sales.create", "orders.view", "orders.create", "crm.view"],
        "retrieve": ["products.view", "sales.view", "sales.create", "orders.view", "orders.create", "crm.view"],
        "create": "products.create",
        "update": "products.edit",
        "partial_update": "products.edit",
        "destroy": "products.delete",
    }

    def get_queryset(self):
        qs = super().get_queryset()
        # Mặc định chỉ trả về sản phẩm đang hoạt động (trừ khi có ?include_inactive=true)
        if self.request.query_params.get("include_inactive") != "true":
            qs = qs.filter(is_active=True)
        # Tìm kiếm theo tên hoặc mã SKU
        search = self.request.query_params.get("search")
        if search:
            qs = qs.filter(name__icontains=search) | qs.filter(sku__icontains=search)
        # Filter theo category
        category_id = self.request.query_params.get("category_id")
        if category_id:
            qs = qs.filter(category_id=category_id)
        return qs

    @action(detail=False, methods=["get"], url_path="export-csv")
    def export_csv(self, request):
        import openpyxl
        from django.http import HttpResponse
        qs = self.get_queryset()
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "SanPham"
        ws.append([
            'Mã SP', 'Tên SP', 'Loại SP', 'Mô tả', 'Đơn vị tính', 'Giá bán', 'Giá nhập', 'Đang kinh doanh'
        ])
        
        for p in qs:
            ws.append([
                p.sku,
                p.name,
                p.category.name if p.category else '',
                p.description,
                p.get_unit_display() if hasattr(p, 'get_unit_display') else p.unit,
                p.price,
                p.cost_price,
                'Có' if p.is_active else 'Không'
            ])
            
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="products.xlsx"'
        wb.save(response)
        return response

    @action(detail=False, methods=["get"], url_path="export-template")
    def export_template(self, request):
        import openpyxl
        from django.http import HttpResponse

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "SanPham_Mau"
        ws.append([
            'Mã SP', 'Tên SP', 'Loại SP', 'Mô tả', 'Đơn vị tính', 'Giá bán', 'Giá nhập'
        ])
        ws.append([
            'SP001', 'Cửa sổ trượt nhôm Xingfa', 'Cửa nhôm', 'Kính cường lực 8mm', 'm2', 1500000, 1200000
        ])
        ws.append([
            'SP002', 'Bản lề 3D', 'Phụ kiện', 'Phụ kiện Kinlong', 'cái', 120000, 100000
        ])
        
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="mau_nhap_san_pham.xlsx"'
        wb.save(response)
        return response

    @action(detail=False, methods=["post"], url_path="import-csv")
    def import_csv(self, request):
        import csv
        import io
        from rest_framework.parsers import MultiPartParser
        
        self.parser_classes = [MultiPartParser]
        file_obj = request.FILES.get('file')
        if not file_obj:
            return Response({"detail": "Vui lòng chọn file."}, status=status.HTTP_400_BAD_REQUEST)
        if not (file_obj.name.endswith('.csv') or file_obj.name.endswith('.xlsx')):
            return Response({"detail": "Chỉ hỗ trợ định dạng .csv hoặc .xlsx"}, status=status.HTTP_400_BAD_REQUEST)
            
        company = request.user.company
        created_count = 0
        updated_count = 0
        
        try:
            if file_obj.name.endswith('.xlsx'):
                import openpyxl
                wb = openpyxl.load_workbook(file_obj, data_only=True)
                ws = wb.active
                rows = list(ws.iter_rows(values_only=True))
                if not rows or len(rows) < 2:
                    return Response({"detail": "File trống."}, status=status.HTTP_400_BAD_REQUEST)
                data_rows = rows[1:]
                
                def get_val(r, idx):
                    if idx < len(r) and r[idx] is not None:
                        return str(r[idx]).strip()
                    return ""
            else:
                decoded_file = file_obj.read().decode('utf-8-sig')
                io_string = io.StringIO(decoded_file)
                reader = csv.reader(io_string)
                headers = next(reader, None)
                if headers and headers[0].startswith('sep='):
                    headers = next(reader, None)
                if not headers:
                    return Response({"detail": "File trống."}, status=status.HTTP_400_BAD_REQUEST)
                data_rows = list(reader)
                
                def get_val(r, idx):
                    if idx < len(r):
                        return str(r[idx]).strip()
                    return ""
                
            for row in data_rows:
                if not row or not any(row): continue
                try:
                    sku = get_val(row, 0)
                    if not sku:
                        continue
                        
                    name = get_val(row, 1)
                    category_name = get_val(row, 2)
                    description = get_val(row, 3)
                    unit = get_val(row, 4).lower() if get_val(row, 4) else "cái"
                    if unit == "mét": unit = "m"
                    elif unit == "lít": unit = "lít"
                    
                    try:
                        price_str = get_val(row, 5)
                        price = float(price_str.replace(',', '')) if price_str else 0
                    except ValueError:
                        price = 0
                        
                    try:
                        cost_str = get_val(row, 6)
                        cost_price = float(cost_str.replace(',', '')) if cost_str else 0
                    except ValueError:
                        cost_price = 0
                        
                    is_active_str = row[7].strip().lower() if len(row) > 7 else "có"
                    is_active = is_active_str not in ["không", "false", "0", "no"]
                    
                    category = None
                    if category_name:
                        category = ProductCategory.objects.filter(company=company, name__iexact=category_name).first()
                        if not category:
                            category = ProductCategory.objects.create(company=company, name=category_name)
                            
                    product = Product.objects.filter(company=company, sku=sku).first()
                    if product:
                        product.name = name or product.name
                        product.category = category
                        product.description = description
                        product.unit = unit
                        product.price = price
                        product.cost_price = cost_price
                        product.is_active = is_active
                        product.save()
                        updated_count += 1
                    else:
                        Product.objects.create(
                            company=company,
                            sku=sku,
                            name=name,
                            category=category,
                            description=description,
                            unit=unit,
                            price=price,
                            cost_price=cost_price,
                            is_active=is_active
                        )
                        created_count += 1
                except Exception as e:
                    continue
                    
        except Exception as e:
            return Response({"detail": f"Lỗi đọc file: {str(e)}"}, status=status.HTTP_400_BAD_REQUEST)
            
        return Response({
            "detail": f"Nhập thành công. Thêm mới: {created_count}, Cập nhật: {updated_count} sản phẩm."
        }, status=status.HTTP_200_OK)


class WarehouseViewSet(TenantQuerySetMixin, viewsets.ModelViewSet):
    """CRUD kho hàng — cô lập theo company."""
    module_code = "inventory"

    queryset = Warehouse.objects.select_related("company").order_by("name")
    serializer_class = WarehouseSerializer
    permission_classes = [permissions.IsAuthenticated, ActionBasedPermission]
    pagination_class = None
    
    action_permissions = {
        "list": ["inventory.manage_warehouse", "inventory.view", "production.manage_factory"],
        "retrieve": ["inventory.manage_warehouse", "inventory.view", "production.manage_factory"],
        "create": "inventory.manage_warehouse",
        "update": "inventory.manage_warehouse",
        "partial_update": "inventory.manage_warehouse",
        "destroy": "inventory.manage_warehouse",
    }

    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        stocks_with_items = StockLevel.objects.filter(warehouse=instance, quantity__gt=0)
        has_txns = InventoryTransaction.objects.filter(warehouse=instance).exists()

        if stocks_with_items.exists() or has_txns:
            target_id = request.query_params.get("target_warehouse_id") or request.data.get("target_warehouse_id")
            if not target_id:
                return Response(
                    {
                        "has_stock": True,
                        "detail": f"Kho hàng '{instance.name}' đang có sản phẩm tồn kho hoặc lịch sử giao dịch. Vui lòng chọn kho nhận để chuyển toàn bộ dữ liệu sang trước khi xoá."
                    },
                    status=status.HTTP_400_BAD_REQUEST
                )

            target_warehouse = Warehouse.objects.filter(company=instance.company, id=target_id).first()
            if not target_warehouse or target_warehouse.id == instance.id:
                return Response(
                    {"detail": "Kho nhận sản phẩm không hợp lệ hoặc trùng với kho đang xoá."},
                    status=status.HTTP_400_BAD_REQUEST
                )

            with transaction.atomic():
                from core.numbering import generate_transaction_code
                for stock in stocks_with_items:
                    target_stock, _ = StockLevel.objects.select_for_update().get_or_create(
                        product=stock.product,
                        warehouse=target_warehouse,
                        defaults={"quantity": 0}
                    )
                    target_stock.quantity += stock.quantity
                    target_stock.save(update_fields=["quantity"])

                    code = generate_transaction_code(instance.company, "adjust")
                    InventoryTransaction.objects.create(
                        company=instance.company,
                        transaction_code=code,
                        product=stock.product,
                        warehouse=target_warehouse,
                        type="adjust",
                        quantity=stock.quantity,
                        note=f"Chuyển tồn kho tự động từ kho đã xoá ({instance.name}) sang ({target_warehouse.name})",
                        created_by=request.user
                    )
                StockLevel.objects.filter(warehouse=instance).delete()
                InventoryTransaction.objects.filter(warehouse=instance).update(warehouse=target_warehouse)
                instance.delete()
            return Response(status=status.HTTP_204_NO_CONTENT)

        StockLevel.objects.filter(warehouse=instance).delete()
        return super().destroy(request, *args, **kwargs)


from rest_framework import mixins

class StockLevelViewSet(mixins.UpdateModelMixin, viewsets.ReadOnlyModelViewSet):
    """Xem và cập nhật (ngưỡng cảnh báo) tồn kho — filter qua product.company."""
    module_code = "inventory"

    queryset = StockLevel.objects.select_related(
        "product__company", "warehouse"
    ).order_by("product__name", "warehouse__name")
    serializer_class = StockLevelSerializer
    permission_classes = [permissions.IsAuthenticated, ActionBasedPermission]
    pagination_class = None
    
    action_permissions = {
        "list": "inventory.view",
        "retrieve": "inventory.view",
        "update": "inventory.manage_warehouse",
        "partial_update": "inventory.manage_warehouse",
    }

    def get_queryset(self):
        user = self.request.user
        if user.is_superuser and user.company_id is None:
            return super().get_queryset()
        qs = self.queryset.filter(product__company=user.company)
        # Filter cảnh báo tồn kho thấp
        if self.request.query_params.get("low_stock") == "true":
            threshold_str = self.request.query_params.get("low_stock_threshold")
            if threshold_str is not None and threshold_str.isdigit():
                threshold = int(threshold_str)
                # Thoả mãn 1 trong 2: bé hơn ngưỡng chung VÀ/HOẶC bé hơn ngưỡng riêng của sản phẩm đó
                low_stock_ids = [s.id for s in qs if s.quantity <= threshold or s.is_low_stock]
            else:
                # Lọc thủ công vì is_low_stock là property
                low_stock_ids = [s.id for s in qs if s.is_low_stock]
            qs = qs.filter(id__in=low_stock_ids)
        warehouse_id = self.request.query_params.get("warehouse_id")
        if warehouse_id:
            qs = qs.filter(warehouse_id=warehouse_id)
        search_query = self.request.query_params.get("search")
        if search_query:
            from django.db.models import Q
            qs = qs.filter(
                Q(product__name__icontains=search_query) |
                Q(product__code__icontains=search_query)
            )
        return qs


class InventoryTransactionViewSet(TenantQuerySetMixin, viewsets.ModelViewSet):
    """
    CRUD phiếu kho — cô lập theo company.
    Backend chặn cứng: type=export chỉ hợp lệ khi đơn hàng đã 'approved'.
    """
    module_code = "inventory"

    queryset = InventoryTransaction.objects.select_related(
        "company", "product", "warehouse", "reference_order", "created_by"
    ).order_by("-created_at")
    serializer_class = InventoryTransactionSerializer
    permission_classes = [permissions.IsAuthenticated, ActionBasedPermission]
    
    action_permissions = {
        "list": "inventory.view",
        "retrieve": "inventory.view",
        "create": ["inventory.import", "inventory.adjust", "inventory.manual_export"],
        "update": "inventory.adjust",
        "partial_update": "inventory.adjust",
    }

    def get_queryset(self):
        qs = super().get_queryset()
        # Filter theo loại phiếu nếu có
        txn_type = self.request.query_params.get("type")
        if txn_type:
            qs = qs.filter(type=txn_type)
        search_query = self.request.query_params.get("search")
        if search_query:
            from django.db.models import Q
            qs = qs.filter(
                Q(transaction_code__icontains=search_query) |
                Q(product__name__icontains=search_query) |
                Q(product__code__icontains=search_query) |
                Q(reference_order__order_number__icontains=search_query) |
                Q(note__icontains=search_query)
            )
        return qs

    def destroy(self, request, *args, **kwargs):
        user_perms = self.request.user.get_permission_codes()
        if not request.user.is_company_admin and not request.user.is_superuser and "inventory.delete_history" not in user_perms:
            return Response(
                {"detail": "Bạn không có quyền xóa lịch sử giao dịch kho."},
                status=status.HTTP_403_FORBIDDEN
            )
        return super().destroy(request, *args, **kwargs)

    @transaction.atomic
    def perform_create(self, serializer):
        from core.numbering import generate_transaction_code
        from orders.models import Order
        from .models import StockLevel

        company = self.request.user.company
        txn_type = serializer.validated_data.get("type")
        reference_order = serializer.validated_data.get("reference_order")
        product = serializer.validated_data.get("product")
        warehouse = serializer.validated_data.get("warehouse")
        quantity = serializer.validated_data.get("quantity", 0)

        # Kiểm tra quyền chi tiết theo loại giao dịch
        from rest_framework.exceptions import PermissionDenied
        user_perms = self.request.user.get_permission_codes()
        if txn_type == "import" and "inventory.import" not in user_perms:
            raise PermissionDenied("Bạn không có quyền nhập kho.")
        if txn_type == "adjust" and "inventory.adjust" not in user_perms:
            raise PermissionDenied("Bạn không có quyền điều chỉnh tồn kho.")
        if txn_type == "transfer" and "inventory.transfer" not in user_perms:
            raise PermissionDenied("Bạn không có quyền điều chuyển kho.")
        if txn_type == "export" and not reference_order and "inventory.manual_export" not in user_perms:
            raise PermissionDenied("Bạn không có quyền tạo xuất kho thủ công.")


        # ─── SECURITY GATE: chặn cứng xuất kho khi đơn chưa duyệt hoặc chưa qua DO Gate ───
        if txn_type == "export" and reference_order:
            if reference_order.status != Order.STATUS_APPROVED:
                from rest_framework import serializers as drf_serializers
                raise drf_serializers.ValidationError(
                    {"reference_order": "Đơn hàng chưa ở trạng thái 'Đã được duyệt', không thể xuất kho."}
                )
            allowed_fin_statuses = [Order.FIN_STATUS_FULLY_PAID, Order.FIN_STATUS_CREDIT_APPROVED]
            if reference_order.financial_status not in allowed_fin_statuses:
                from rest_framework import serializers as drf_serializers
                raise drf_serializers.ValidationError(
                    {"reference_order": f"Khóa Xuất Kho (DO Gate): Đơn {reference_order.order_number} chưa thanh toán đủ hoặc chưa được phê duyệt xuất kho nợ."}
                )

        # Sinh mã phiếu tự động
        transaction_code = generate_transaction_code(company, txn_type)
        serializer.save(
            company=company,
            created_by=self.request.user,
            transaction_code=transaction_code,
        )

        # ─── Cập nhật tồn kho (StockLevel) ───
        if warehouse and product:
            stock, _ = StockLevel.objects.select_for_update().get_or_create(
                product=product,
                warehouse=warehouse,
                defaults={"quantity": 0},
            )
            if txn_type == "import":
                stock.quantity += quantity
                stock.save(update_fields=["quantity"])
            elif txn_type == "adjust":
                stock.quantity = quantity
                stock.save(update_fields=["quantity"])
            elif txn_type == "export" and not reference_order:
                stock.quantity = max(0, stock.quantity - quantity)
                stock.save(update_fields=["quantity"])
            elif txn_type == "transfer":
                stock.quantity = max(0, stock.quantity - quantity)
                stock.save(update_fields=["quantity"])
                
                # Cộng tồn kho cho target_warehouse
                target_warehouse = serializer.validated_data.get("target_warehouse")
                if target_warehouse:
                    target_stock, _ = StockLevel.objects.select_for_update().get_or_create(
                        product=product,
                        warehouse=target_warehouse,
                        defaults={"quantity": 0},
                    )
                    target_stock.quantity += quantity
                    target_stock.save(update_fields=["quantity"])

    @action(detail=True, methods=["post"], permission_classes=[permissions.IsAuthenticated, ActionBasedPermission])
    @transaction.atomic
    def approve(self, request, pk=None):
        """Duyệt lệnh xuất kho và trừ tồn kho"""
        txn = self.get_object()
        
        # Check permissions
        if not request.user.is_company_admin and not request.user.role.permissions.filter(code="inventory.approve_export").exists():
            return Response({"detail": "Bạn không có quyền duyệt lệnh xuất kho."}, status=status.HTTP_403_FORBIDDEN)

        if txn.type != "export" or txn.status != txn.STATUS_PENDING:
            return Response({"detail": "Chỉ có thể duyệt các lệnh xuất kho đang ở trạng thái chờ duyệt."}, status=status.HTTP_400_BAD_REQUEST)
            
        warehouse_id = request.data.get("warehouse_id")
        if not warehouse_id:
            return Response({"detail": "Vui lòng chọn kho để xuất."}, status=status.HTTP_400_BAD_REQUEST)
            
        warehouse = Warehouse.objects.filter(company=txn.company, id=warehouse_id).first()
        if not warehouse:
            return Response({"detail": "Kho hàng không hợp lệ."}, status=status.HTTP_400_BAD_REQUEST)

        # Check and update stock
        from .models import StockLevel
        stock, _ = StockLevel.objects.select_for_update().get_or_create(
            product=txn.product,
            warehouse=warehouse,
            defaults={"quantity": 0}
        )
        
        if stock.quantity < txn.quantity:
            return Response({
                "detail": f"Không đủ tồn kho! Sản phẩm '{txn.product.sku}' trong kho '{warehouse.name}' chỉ còn {stock.quantity} (cần xuất {txn.quantity}).",
                "stock_quantity": stock.quantity
            }, status=status.HTTP_400_BAD_REQUEST)
            
        # Deduct stock
        stock.quantity -= txn.quantity
        stock.save(update_fields=["quantity"])
        
        # Cảnh báo tồn kho thấp
        if stock.is_low_stock and stock.min_quantity > 0:
            try:
                from notifications.utils import notify_inventory_low
                notify_inventory_low(stock)
            except Exception:
                pass
                
        # Update transaction
        factory_id = request.data.get("factory_id")
        factory_obj = None
        if factory_id:
            from production.models import Factory
            factory_obj = Factory.objects.filter(id=factory_id).first()
        
        txn.warehouse = warehouse
        txn.status = txn.STATUS_COMPLETED
        txn.factory = factory_obj
        txn.save(update_fields=["warehouse", "status", "factory"])
        
        # Kích hoạt Lệnh Sản Xuất cho Đơn hàng
        if txn.reference_order and factory_obj:
            try:
                from orders.signals import _create_production_order
                _create_production_order(txn.reference_order, factory_id=factory_id)
            except Exception as e:
                import logging
                logging.getLogger(__name__).error(f"Lỗi khi auto-create MO từ kho: {e}")
        
        return Response({"detail": "Đã duyệt và xuất kho thành công."})

    @action(detail=True, methods=["post"], permission_classes=[permissions.IsAuthenticated, ActionBasedPermission])
    def reject(self, request, pk=None):
        """Từ chối/hủy lệnh xuất kho"""
        txn = self.get_object()
        
        # Check permissions
        if not request.user.is_company_admin and not request.user.role.permissions.filter(code="inventory.approve_export").exists():
            return Response({"detail": "Bạn không có quyền từ chối lệnh xuất kho."}, status=status.HTTP_403_FORBIDDEN)

        if txn.type != "export" or txn.status != txn.STATUS_PENDING:
            return Response({"detail": "Chỉ có thể từ chối các lệnh xuất kho đang ở trạng thái chờ duyệt."}, status=status.HTTP_400_BAD_REQUEST)
            
        txn.status = txn.STATUS_REJECTED
        txn.save(update_fields=["status"])
        
        # Nếu phiếu xuất thuộc về 1 Đơn hàng đã được Duyệt, giữ nguyên trạng thái 'approved'
        # (Đơn vẫn được giám đốc duyệt, chỉ là kho chưa xuất được do thiếu hàng/cần điều chuyển)
        # Sale bây giờ có thể bấm "Üyêu cầu xuất kho lại" để thủ kho duyệt lại
        if txn.reference_order and txn.reference_order.status == 'approved':
            order = txn.reference_order
            # Hủy ApprovalRequest đang chờ (nếu có) của đơn nợ (finance.approve_credit)
            # nhưng KHAI KHAI đồng đạm trạng thái đơn hàng
            try:
                from approvals.models import ApprovalRequest
                from django.contrib.contenttypes.models import ContentType
                ct = ContentType.objects.get_for_model(order)
                ApprovalRequest.objects.filter(
                    content_type=ct,
                    object_id=order.id,
                    status=ApprovalRequest.STATUS_PENDING,
                    title__startswith="Duyệt xuất kho nợ"
                ).update(status=ApprovalRequest.STATUS_CANCELLED)
            except Exception:
                pass

        return Response({"detail": "Đã hủy lệnh xuất kho. Sale có thể yêu cầu xuất kho lại."})

    @action(detail=True, methods=["post"], url_path="recreate-production-order",
            permission_classes=[permissions.IsAuthenticated, ActionBasedPermission])
    def recreate_production_order(self, request, pk=None):
        """
        Tạo lại Lệnh Sản Xuất cho phiếu xuất kho đã hoàn thành
        khi lệnh SX bị xóa mất hoặc chưa được tạo.
        Yêu cầu quyền: production.create hoặc Company Admin.
        """
        txn = self.get_object()

        # Kiểm tra quyền: phải là Company Admin hoặc có quyền tạo lệnh SX
        user = request.user
        if not user.is_company_admin and not user.is_superuser:
            user_perms = user.get_permission_codes()
            if "production.create" not in user_perms:
                return Response(
                    {"detail": "Bạn không có quyền tạo lệnh sản xuất."},
                    status=status.HTTP_403_FORBIDDEN
                )

        # Chỉ áp dụng cho phiếu xuất đã hoàn thành và có liên kết đơn hàng
        if txn.type != InventoryTransaction.TYPE_EXPORT:
            return Response({"detail": "Chỉ có thể tạo lại lệnh SX từ phiếu xuất kho."}, status=status.HTTP_400_BAD_REQUEST)
        if txn.status != InventoryTransaction.STATUS_COMPLETED:
            return Response({"detail": "Phiếu xuất kho chưa hoàn thành, không thể tạo lệnh SX."}, status=status.HTTP_400_BAD_REQUEST)
        if not txn.reference_order:
            return Response({"detail": "Phiếu xuất kho này không liên kết với đơn hàng nào."}, status=status.HTTP_400_BAD_REQUEST)

        order = txn.reference_order
        # Kiểm tra đã có lệnh SX chưa
        if order.production_orders.exists():
            return Response({"detail": "Lệnh sản xuất đã tồn tại cho đơn hàng này."}, status=status.HTTP_400_BAD_REQUEST)

        try:
            from orders.signals import _create_production_order
            factory_id = txn.factory_id  # Lấy từ phiếu xuất (nếu có gắn nhà máy)
            _create_production_order(order, factory_id=factory_id)
            # Lấy mã lệnh SX vừa tạo để trả về
            po = order.production_orders.first()
            return Response({
                "detail": f"Đã tạo lại Lệnh Sản Xuất thành công.",
                "production_order_code": po.production_order_code if po else None,
            })
        except Exception as e:
            import logging
            logging.getLogger(__name__).error(f"Lỗi tạo lại ProductionOrder từ kho: {e}")
            return Response({"detail": f"Lỗi khi tạo lại lệnh SX: {str(e)}"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=False, methods=["delete"], url_path="clear-history")
    def clear_history(self, request):
        user_perms = self.request.user.get_permission_codes()
        if not request.user.is_company_admin and not request.user.is_superuser and "inventory.delete_history" not in user_perms:
            return Response(
                {"detail": "Bạn không có quyền xóa lịch sử giao dịch kho."},
                status=status.HTTP_403_FORBIDDEN
            )
        count, _ = self.get_queryset().delete()
        return Response({"detail": f"Đã xóa {count} giao dịch kho thành công."})

    @action(detail=False, methods=["delete"], url_path="delete-by-code")
    def delete_by_code(self, request):
        user_perms = self.request.user.get_permission_codes()
        if not request.user.is_company_admin and not request.user.is_superuser and "inventory.delete_history" not in user_perms:
            return Response(
                {"detail": "Bạn không có quyền xóa lịch sử giao dịch kho."},
                status=status.HTTP_403_FORBIDDEN
            )
        code = request.query_params.get("code")
        if not code:
            return Response({"detail": "Thiếu mã phiếu (transaction_code)."}, status=status.HTTP_400_BAD_REQUEST)
            
        qs = self.get_queryset().filter(transaction_code=code)
        if not qs.exists():
            return Response({"detail": "Không tìm thấy giao dịch nào với mã này."}, status=status.HTTP_404_NOT_FOUND)
            
        count, _ = qs.delete()
        return Response({"detail": f"Đã xóa {count} giao dịch thuộc phiếu {code}."})

    @action(detail=False, methods=["post"], url_path="bulk-create")
    @transaction.atomic
    def bulk_create(self, request):
        """Tạo giao dịch kho hàng loạt (cùng 1 mã phiếu)."""
        from core.numbering import generate_transaction_code
        from orders.models import Order
        from .models import StockLevel, InventoryTransaction, Product, Warehouse

        company = request.user.company
        data = request.data
        txn_type = data.get("type")
        warehouse_id = data.get("warehouse")
        target_warehouse_id = data.get("target_warehouse")
        note = data.get("note", "")
        items = data.get("items", [])

        if not items or not isinstance(items, list):
            return Response({"detail": "Danh sách sản phẩm không hợp lệ."}, status=status.HTTP_400_BAD_REQUEST)

        # Kiểm tra quyền chi tiết theo loại giao dịch
        from rest_framework.exceptions import PermissionDenied
        user_perms = self.request.user.get_permission_codes()
        if txn_type == "import" and "inventory.import" not in user_perms:
            raise PermissionDenied("Bạn không có quyền nhập kho.")
        if txn_type == "adjust" and "inventory.adjust" not in user_perms:
            raise PermissionDenied("Bạn không có quyền điều chỉnh tồn kho.")
        if txn_type == "transfer" and "inventory.transfer" not in user_perms:
            raise PermissionDenied("Bạn không có quyền điều chuyển kho.")
        if txn_type == "export" and "inventory.manual_export" not in user_perms:
            raise PermissionDenied("Bạn không có quyền tạo xuất kho thủ công.")

        # Lấy thông tin kho
        warehouse = Warehouse.objects.filter(id=warehouse_id, company=company).first() if warehouse_id else None
        target_warehouse = Warehouse.objects.filter(id=target_warehouse_id, company=company).first() if target_warehouse_id else None

        if not warehouse:
            return Response({"warehouse": "Vui lòng chọn Kho xuất."}, status=status.HTTP_400_BAD_REQUEST)

        if txn_type == "transfer":
            if not target_warehouse:
                return Response({"target_warehouse": "Vui lòng chọn Kho nhận."}, status=status.HTTP_400_BAD_REQUEST)
            if warehouse.id == target_warehouse.id:
                return Response({"target_warehouse": "Kho xuất và Kho nhận không được trùng nhau."}, status=status.HTTP_400_BAD_REQUEST)

        # Sinh mã phiếu tự động
        transaction_code = generate_transaction_code(company, txn_type)

        created_transactions = []
        for index, item in enumerate(items):
            product_id = item.get("product")
            quantity = int(item.get("quantity", 0))
            unit_cost = item.get("unit_cost", 0)

            product = Product.objects.filter(id=product_id, company=company).first()
            if not product:
                return Response({"detail": f"Sản phẩm ID {product_id} không tồn tại."}, status=status.HTTP_400_BAD_REQUEST)

            if quantity <= 0:
                return Response({"detail": f"Số lượng sản phẩm '{product.name}' phải lớn hơn 0."}, status=status.HTTP_400_BAD_REQUEST)

            # Khóa và lấy stock hiện tại
            stock, _ = StockLevel.objects.select_for_update().get_or_create(
                product=product,
                warehouse=warehouse,
                defaults={"quantity": 0},
            )

            # Kiểm tra tồn kho cho export và transfer
            if txn_type in ["export", "transfer"] and stock.quantity < quantity:
                return Response(
                    {"detail": f"Sản phẩm '{product.name}' không đủ tồn kho! Tồn kho hiện tại: {stock.quantity}"},
                    status=status.HTTP_400_BAD_REQUEST
                )

            # Cập nhật StockLevel
            if txn_type == "import":
                stock.quantity += quantity
            elif txn_type == "adjust":
                stock.quantity = quantity
            elif txn_type in ["export", "transfer"]:
                stock.quantity -= quantity
            
            stock.save(update_fields=["quantity"])

            if txn_type == "transfer":
                target_stock, _ = StockLevel.objects.select_for_update().get_or_create(
                    product=product,
                    warehouse=target_warehouse,
                    defaults={"quantity": 0},
                )
                target_stock.quantity += quantity
                target_stock.save(update_fields=["quantity"])

            # Tạo record
            txn = InventoryTransaction(
                company=company,
                transaction_code=transaction_code,
                type=txn_type,
                status=InventoryTransaction.STATUS_COMPLETED,
                product=product,
                warehouse=warehouse,
                target_warehouse=target_warehouse if txn_type == "transfer" else None,
                quantity=quantity,
                unit_cost=unit_cost,
                note=note,
                created_by=request.user
            )
            txn.save()
            created_transactions.append(txn)

        serializer = self.get_serializer(created_transactions, many=True)
        return Response(serializer.data, status=status.HTTP_201_CREATED)
